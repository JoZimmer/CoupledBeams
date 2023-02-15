import pickle
import numpy as np
import os
from os.path import join as os_join
from os.path import sep as os_sep
import string 
import xlwings as xl
import pandas as pd
import copy

# from source.utilities import statistics_utilities as stats_utils
from source.utilities import global_definitions as GD

#_____________ ALTES VOM URSPRÜNGLICHEN BEAM _____________________

def evaluate_residual(a_cur, a_tar):
    residual = np.linalg.norm(np.subtract(a_cur, a_tar)) / np.amax(np.absolute(a_tar))

    # print ('current: ', a_cur)
    # print ('target: ', a_tar)
    # print('residual:', residual)
    # print()
    return residual

def cm2inch(value):
    return value/2.54

def check_and_flip_sign_dict(eigenmodes_dict):
    '''
    flips the sign of y and a deformation of modes to be positive at the first node after ground 
    dependend/coupled dofs are flip accordingly
    '''

    for idx in range(len(eigenmodes_dict['y'])):
        if eigenmodes_dict['y'][idx][1] < 0:
            eigenmodes_dict['y'][idx] = np.negative(eigenmodes_dict['y'][idx])
            try:
                eigenmodes_dict['g'][idx] = np.negative(eigenmodes_dict['g'][idx])
            except KeyError:
                pass
        try:
            if eigenmodes_dict['a'][idx][1] < 0:
                eigenmodes_dict['a'][idx] = np.negative(eigenmodes_dict['a'][idx])
        except KeyError:
            pass

    return eigenmodes_dict

def check_and_flip_sign_array(mode_shape_array):
    '''
    check_and_change_sign
    change the sign of the mode shape such that the first entry is positive
    '''
    if mode_shape_array[1] < 0:
        return mode_shape_array * -1
    elif mode_shape_array [1] > 0:
        return mode_shape_array

def analytic_function_static_disp(parameters, x, load_type = 'single'):
    l = parameters['lz_total_beam']
    EI = parameters['E_Modul'] * parameters['Iz']
    magnitude = parameters['static_load_magnitude']
    if load_type == 'single':
        #print ('  w_max soll:', magnitude*l**3/(3*EI))
        return (magnitude/EI) * (0.5 * l * (x**2) - (x**3)/6) 
    elif load_type == 'continous':
        #print ('  w_max soll:', magnitude*l**4/(8*EI))
        return -(magnitude/EI) * (-x**4/24 + l * x**3 /6 - l**2 * x**2 /4)

def analytic_eigenfrequencies(beam_model):
    # von https://me-lrt.de/eigenfrequenzen-eigenformen-beim-balken 
    parameters = beam_model.parameters
    l = parameters['lz_total_beam']
    EI = parameters['E_Modul'] * parameters['Iz']
    A = parameters['cross_section_area']
    rho = parameters['material_density']
    lambdas = [1.875, 4.694, 7.855]
    f_j = np.zeros(3)
    for i, l_i in enumerate(lambdas):
        f_j[i] = np.sqrt((l_i**4 * EI)/(l**4 * rho *A)) / (2*np.pi)
    
    return f_j

def analytic_eigenmode_shapes(beam_model):
    # von https://me-lrt.de/eigenfrequenzen-eigenformen-beim-balken 
    #parameters = beam_model.parameters
    l = beam_model.parameters['lz_total_beam']
    x = beam_model.nodal_coordinates['x0']
    lambdas = [1.875, 4.694, 7.855] # could also be computed as seen in the link

    w = []
    for j in range(3):
        zeta = lambdas[j] * x / l
        a = (np.sin(lambdas[j]) - np.sinh(lambdas[j])) / (np.cos(lambdas[j]) + np.cosh(lambdas[j])) 
        
        w_j = np.cos(zeta) - np.cosh(zeta) + a*(np.sin(zeta) -np.sinh(zeta))
        w.append(w_j)
        
        reduced_m = beam_model.m[::2,::2] 
        gen_mass = np.dot(np.dot(w_j, reduced_m), w_j)
        norm_fac = np.sqrt(gen_mass)
        w_j /= norm_fac
        is_unity = np.dot(np.dot(w_j, reduced_m), w_j)
        if round(is_unity, 4) != 1.0:
            raise Exception ('analytic mode shape normalization failed')

    return w

def save_optimized_beam_parameters(opt_beam_model, fname):
    import json
    new = 'optimized_parameters'+os_sep+fname+'.json'
    if os.path.isfile(new):
        print('WARNING', new, 'already exists!')
        new = 'optimized_parameters'+os_sep+fname+'_1.json'

    f = open(new,'w')
    
    json.dump(opt_beam_model.parameters, f)
    f.close()
    print('\nsaved:', new)

def remove_small_values_from_array(arr, tolerance=1e-05):

    for i, val in enumerate(arr):
        if abs(val) <= tolerance:
            arr[i] = 0
    return arr

#________________ LASTEN ____________________________________

def parse_load_signal(signal_raw, time_array=None, load_directions_to_include = 'all', discard_time = None, dimension = '2D'):
    '''
    - sorts the load signals in a dictionary with load direction as keys:
        x,y,z: nodal force
        a,b,g: nodal moments
    - deletes first entries until discard_time
    - only gives back the components specified, default is all 
    TODO: instead of setting unused direction to 0 do some small number or exclude them differntly 
    '''
    if discard_time:
        signal_raw = signal_raw[:,discard_time:]
    if load_directions_to_include == 'all':
        load_directions_to_include = ['Fx', 'Fy', 'Mx', 'Mz']
    n_nodes = int(signal_raw.shape[0]/GD.DOFS_PER_NODE[dimension])
    
    signal = {}
    for i, label in enumerate(GD.DOF_LABELS[dimension]):
        if GD.DOF_LOAD_MAP[label] in load_directions_to_include:
            signal[label] = signal_raw[i::GD.DOFS_PER_NODE[dimension]]
        else:
            signal[label] = np.zeros((n_nodes,signal_raw.shape[1]))

    # if time_array.any():
    #     dt = time_array[1] - time_array[0] # simulation time step
    # else:
    #     dt = 0.1 # some default

    # signal['sample_freq'] = 1/dt

    return signal

def parse_load_signal_backwards(signal, domain_size):
    '''
    signal kommt als dictionary mit den Richtungen als keys (müssen nicht alle 6 Richtugne sein)
    output soll row vector sein mit dofs * n_nodes einträgen
    '''
    from source.utilities import global_definitions as GD

    shape = GD.DOFS_PER_NODE[domain_size] * len(list(signal.values())[0])
    signal_raw = np.zeros(shape)
    for label in signal:
        dof_label_id = GD.DOF_LABELS[domain_size].index(label)
        for i, val in enumerate(signal[label]):
            sort_id = i * GD.DOFS_PER_NODE[domain_size] + dof_label_id
            signal_raw[sort_id] = val

    #TODO: must this be transposed?!
    return signal_raw

def generate_static_load_vector_file(load_vector):
    '''
    return a npy file and saves it for a given load.

    If load is given as dictionary with directions as keys it parses it to a 1D array that is needed in the ParOptBeam
    '''
    if isinstance(load_vector, dict):
        load = parse_load_signal_backwards(load_vector)
    else:
        load = load_vector

    src_path = os_join('input','loads','static')
    if not os.path.isdir(src_path):
        os.makedirs(src_path)

    file_name = src_path + 'eswl_' + str(int(len(load)/GD.DOFS_PER_NODE['3D'])) + '_nodes.npy'

    np.save(file_name, load)

    return file_name

def generate_nodal_force_file(number_of_nodes, node_of_load_application, force_direction:str, magnitude, domain_size = '3D', file_base_name = 'static_load_'):
    '''
    creating a force .npy file with a nodal force at given node, direction and magnitude
    number_of_nodes: anzahl knoten des models 
    node_of_load_application: Knoten an dem die Last aufgebracht werden soll
    force_direction: string der richtung
    Einheit der Kraft = [N]
    '''
    from source.utilities import global_definitions as GD
    src_path = os_join(*['inputs','loads','static'])
    if not os.path.isdir(src_path):
        os.makedirs(src_path)

    force_file_name = os_join(src_path, file_base_name + str(number_of_nodes) + '_nodes_at_' + str(node_of_load_application) + \
                        '_in_' + force_direction+'.npy')

    if os.path.isfile(force_file_name):
        return force_file_name
    
    else:
        # -1 ist notwendig da sozusagen vom knoten unter dem von interesse angefangen wird die dof nodes dazu zu addieren
        loaded_dof = (node_of_load_application-1)*GD.DOFS_PER_NODE[domain_size] + GD.DOF_LABELS[domain_size].index(force_direction)

        force_data = np.zeros(GD.DOFS_PER_NODE[domain_size]*number_of_nodes)
        force_data[loaded_dof] += magnitude
        force_data = force_data.reshape(GD.DOFS_PER_NODE[domain_size]*number_of_nodes,1)

        np.save(force_file_name, force_data)

        return force_file_name

def generate_kopflasten_file(number_of_nodes, loads_dict, file_base_name = 'kopflasten_'):
    '''
    creating a force .npy file with a nodal force at given node, direction and magnitude
    number_of_nodes: anzahl knoten des models 
    node_of_load_application: Knoten an dem die Last aufgebracht werden soll
    force_direction: string der richtung
    Einheit der Kraft = [N]
    '''
    from source.utilities import global_definitions as GD
    src_path = os_join(*['inputs','loads','static'])
    if not os.path.isdir(src_path):
        os.makedirs(src_path)

    node_of_load_application = number_of_nodes
    force_file_name = os_join(src_path, file_base_name + '.npy')

    # if os.path.isfile(force_file_name):
    #     return force_file_name
    
    #else:
    # -1 ist notwendig da sozusagen vom knoten unter dem von interesse angefangen wird die dof nodes dazu zu addieren
    force_data = np.zeros(GD.DOFS_PER_NODE['2D']*number_of_nodes)
    for load_direction, magnitude in loads_dict.items():

        force_direction = GD.LOAD_DOF_MAP[load_direction]
        loaded_dof = (node_of_load_application-1)*GD.DOFS_PER_NODE['2D'] + GD.DOF_LABELS['2D'].index(force_direction)
        force_data[loaded_dof] += magnitude

    force_data = force_data.reshape(GD.DOFS_PER_NODE['2D']*number_of_nodes,1)

    np.save(force_file_name, force_data)

    return force_file_name

def horizontale_ersatzlast(vertikale_lasten:list, theta_x):
    '''
    Bisher annahme das Fz Lasten immer nach unten wirken stimmt in allen bisherigen Fällen aber ist ansich nciht richtig für schiefstellung
    vertikale_lasten: lsite an dict die unter 'Fz' die vertiakeln Lasten enthalten die einen Effekt bei der Scheifstellung haben
    theta_x: schiefstellung in m/m 
    '''

    F_h_ers = {'Fx':np.zeros(vertikale_lasten[0]['Fz'].shape)}

    for last in vertikale_lasten:

        F_h_ers['Fx'] += last['Fz'] * theta_x * -1

    return F_h_ers

def update_lasten_dict(lasten_dict_base, lasten_dicts_to_add:list, wind_kraft=None, kopflasten=None, abminderung_wind = 1.0):
    '''
    merged die Kopflasten und die windkraft nach DIN 
    lasten_dict_base: gibt größe und bezeichnugn der kräfte vor
    wenn windkraft = None wird nur die Kopflast in das richtige format gebracht
    abminderung_wind: im Bauzustand = 0.5?!
    '''
    lasten_dict_updated = copy.deepcopy(lasten_dict_base)

    # if isinstance(wind_kraft, np.ndarray):
    #     lasten_dict_updated['Fy'] += wind_kraft
    # for i, direction in enumerate(lasten_dict_base):
    #     lasten_dict_updated[direction][-1] += kopflasten[direction]

    for i, last in enumerate(lasten_dicts_to_add):
        for direction in lasten_dict_base:
            if direction not in last.keys():
                continue
            lasten_dict_updated[direction] += last[direction]
            if i == 0 and abminderung_wind != 1.0:
                lasten_dict_updated[direction] *= abminderung_wind

    return lasten_dict_updated

def lasten_dict_bauzustand(lasten_dict_base, knoten_wind_kraft_z, gewichtskraft_charkteristisch, sicherheitsbeiwerte, QS_obj, abminderung_wind, parameter):
    '''
    lasten_dict_end: Lasten dict des Turms über die gesamte Höhe
    '''
    # Lasten wenn alle Segmente drauf sind
    lasten_dict_base_ebenen = {k: np.zeros(QS_obj.n_ebenen) for k in lasten_dict_base}
    lasten_dict_end = update_lasten_dict(lasten_dict_base_ebenen, [knoten_wind_kraft_z['Ebenen'], gewichtskraft_charkteristisch], abminderung_wind=abminderung_wind)
    lasten_bauzustand = {}

    # Das Eigengewicht wird nur am unteren Knoten aufgebracht und nicht oben schon 
    knoten_komponente = {'Fz':0, 'Fx':1, 'Fy':1, 'Mx':1,'My':1, 'Mz':1}
    for segment in range(1,QS_obj.n_ebenen):
        seg = 'seg_0-' + str(segment)
        h_ebene = QS_obj.section_absolute_heights['Ebenen'][segment]
        relevante_knoten = [j for j, h  in enumerate(QS_obj.section_absolute_heights['Ebenen']) if h <= h_ebene]
        lasten_bauzustand[seg] = {}
        for komponente in lasten_dict_end:
            modfier = np.zeros(len(lasten_dict_end[komponente]))
            modfier[:relevante_knoten[-1]+knoten_komponente[komponente]] = 1.
            lasten_bauzustand[seg][komponente] = lasten_dict_end[komponente] * modfier

        # SCHIEFSTELLUNG -> hierfür wird die Gewichtskraft wieder mit Sicherheitsbeiwert berücksichtigt
        lasten_bauzustand[seg]['Fx'] += lasten_bauzustand[seg]['Fz'] * parameter['imperfektion'] * sicherheitsbeiwerte['g']

    return lasten_bauzustand

def generate_lasten_file(number_of_nodes, lasten_dict, file_base_name, dimension:str = '2D'):
    '''
    lasten kommen als dictionary mit den Richtungen als keys (siehe GD.LOAD_DOF_MAP)
    und werden in einen 1D vektor umgewandelt 
    return: datei name da last als npy datei gespeichert wird
    dimension 2D o. 3D
    '''

    src_path = os_join(*['inputs','loads','static'])
    if not os.path.isdir(src_path):
        os.makedirs(src_path)

    force_file_name = os_join(src_path, file_base_name + '.npy')

    force_data = np.zeros(GD.DOFS_PER_NODE[dimension]*number_of_nodes)
    for load_direction, magnitude in lasten_dict.items():
        if dimension == '2D' and load_direction == 'Mz':
            continue
        force_direction = GD.LOAD_DOF_MAP[load_direction]
        start = GD.DOF_LABELS[dimension].index(force_direction)
        step = len(GD.DOF_LABELS[dimension])

        force_data[start::step] += magnitude

    force_data = force_data.reshape(GD.DOFS_PER_NODE[dimension]*number_of_nodes,1)

    np.save(force_file_name, force_data)

    return force_file_name

def linien_last_to_knoten_last(last_vektor, knoten_z, gamma_q=1.5):

    '''
    vekotr der eine Last pro meter darstellt in knotenlasten vektor überführen 
    '''
    if not isinstance(knoten_z, np.ndarray):
        knoten_z = np.array(knoten_z)

    # Hebelarme/zu integrierende Fläche
    d_i = np.diff(knoten_z)
            
    d_i[0] = d_i[1]/2
    d_i = np.append(d_i, d_i[1]/2)    

    if isinstance(last_vektor, dict):
        F_z = {}
        for last_key in last_vektor:
            F_z[last_key] = gamma_q * last_vektor[last_key] * np.array(d_i)
    
    else:
        F_z = gamma_q * last_vektor * np.array(d_i)

    return F_z

def parse_schnittgrößen_labels(lasten_dict):
    '''
    lasten dict -> tiefe über knoten_typ, QS_label, nabenhöhe, dauer
    '''
    lasten_dict_parsed = copy.deepcopy(lasten_dict)
    for knoten in lasten_dict:
        for QS_label in lasten_dict[knoten]:
            for nabenhöhe in lasten_dict[knoten][QS_label]:
                for dauer, lasten in lasten_dict[knoten][QS_label][nabenhöhe].items():
                    for direction in lasten:
                        lasten_dict_parsed[knoten][QS_label][nabenhöhe][dauer][GD.DOF_RESPONSE_MAP[direction]] = lasten[direction]
                        del lasten_dict_parsed[knoten][QS_label][nabenhöhe][dauer][direction]

    return lasten_dict_parsed

def interpolate(z_soll , z_ist, werte):
    
    interpolierte_werte = []
    for z in z_soll:
        for i, z_i in enumerate(z_ist[:-1]):
            z1, z2 = z_ist[i], z_ist[i+1]
            if z1 <= z < z2:
                y1 = werte[i]
                y2 = werte[i+1]
                val_z = np.interp(z, (z1,z2), (y1,y2))
                interpolierte_werte.append(val_z)
    interpolierte_werte.append(werte[-1])

    return np.asarray(interpolierte_werte)

def kragarm_einflusslinien(nodal_coordinates, load_direction, node_id, response, response_node_id=0):
    '''
    for a lever arm this is simple
    if shear response -> return 1
    if base moment -> return level* 1
    '''
    moment_load = {'x':'My', 'y':'Mx', 'g':'Mz','a':'Mx', 'b':'My'}
    shear_load = {'y':'Qy', 'z':'Qz'}

    #nodal_coordinates = structure_model.nodal_coordinates['x0']

    if load_direction == 'x':
        if moment_load[load_direction] == response:
            # positive
            if node_id - response_node_id <= 0:
                return 0.0
            else:
                return nodal_coordinates[node_id - response_node_id]

        elif shear_load[load_direction] == response:
            if node_id >= response_node_id:
                return 1.0
            else:
                return 0.0
        else:
            return 0.0

    elif load_direction == 'y':
        if moment_load[load_direction] == response:
            # negative
            if node_id - response_node_id <= 0:
                return 0.0
            else:
                return -nodal_coordinates[node_id - response_node_id]
        elif shear_load[load_direction] == response:
            if node_id >= response_node_id:
                return 1.0
            else:
                return 0.0

        else:
            return 0.0

    elif load_direction == 'z':
        return 0.0

    elif load_direction in ['a','b','g']:
        unit = '[Nm]'
        if moment_load[load_direction] == response:
            if node_id >= response_node_id:
                return 1.0
            else:
                return 0.0
        else: # moments don't cause shear forces
            return 0.0

#____________________DYNAMIC ANALYSIS________________________________
def get_fft(given_series, sampling_freq):

    '''
    The function get_fft estimates the Fast Fourier transform of the given signal 
    sampling_freq = 1/dt
    '''

    signal_length=len(given_series)

    freq_half =  np.arange(0, 
                           sampling_freq/2 - sampling_freq/signal_length + sampling_freq/signal_length, 
                           sampling_freq/signal_length)

    # single sided fourier
    series_fft = np.fft.fft(given_series)
    series_fft = np.abs(series_fft[0:int(np.floor(signal_length/2))])/np.floor(signal_length/2)  
    
    max_length = len(freq_half)
    if max_length < len(series_fft):
        max_length = len(series_fft)
    
    freq_half = freq_half[:max_length-1]
    series_fft = series_fft[:max_length-1]
    
    return freq_half, series_fft

def extreme_value_analysis_nist(given_series, dt, response_label = None, type_of_return = 'estimate', P1 = 0.98):
    ''' 
    dynamic_analysi_solved: dynamic_analysis.solver object
    response: label given as dof_label, if given as response label it is convertedd
    type_of_return: wheter the estimated or the quantile value of P1 is returned (both are computed)
    ''' 

    T_series = dt * len(given_series)
    dur_ratio = 600 / T_series
    # # MAXMINEST NIST
    #P1 = 0.98
    max_qnt, min_qnt, max_est, min_est, max_std, min_std, Nupcross = stats_utils.maxmin_qnt_est(given_series	, 
                                                                        cdf_p_max = P1 , cdf_p_min = 0.0001, cdf_qnt = P1, dur_ratio = dur_ratio)
    
    abs_max_est = max([abs(max_est[0][0]), abs(min_est[0][0])])
    abs_max_qnt = max([abs(max_qnt[0]), abs(min_qnt[0])])

    if type_of_return == 'estimate':
        extreme_response = abs_max_est
    elif type_of_return == 'quantile':
        extreme_response = abs_max_qnt
    
    glob_max = max(abs(given_series))

    return abs_max_qnt, abs_max_est

#______________________SONSTIGES_____________________________________

def add_defaults(params_dict:dict):
    '''
    Im run sind nur die tatsächlich relevanten werte zu definieren. Weitere Werte die aufgrund von älteren Versionen gebruacht werden werden dann hier ergäntz
    '''
    parameters = {
                'lz_total_beam': 130, # wird mit der Querschnitts definition ergänzt
                'material_density': 904,# für dynamische Berechnung äquivalenter Wert - nicht relevant
                'total_mass_tower': 668390,# aus RFEM - nicht relevant
                'E_Modul': 12000E+06,# N/m²
                'nu': 0.1, # querdehnung
                'damping_coeff': 0.025,
                'nodes_per_elem': 2,
                'Iz': 51.0,# Not used if defined on intervals
                'dofs_of_bc':[0,1,2], # Einspannung
                'dynamic_load_file': os_join(*["inputs","loads","dynamic","dynamic_force_11_nodes.npy"]),
                'time_step_dynamic_load':0.01, # Time step der simulation
                'eigen_freqs_target':[0.133,0.79,3.0], 
            }
    params_dict.update(parameters)

    einspannung = {'2D':[0,1,2], '3D':[0,1,2,3,4,5]}
    params_dict['dofs_of_bc'] = einspannung[params_dict['dimension']]

    return params_dict

def load_object_from_pkl(pkl_file):
    '''full path mit datei name zum jobejt file'''
    with open(pkl_file, 'rb') as handle:
        data = pickle.load(handle)

    return data

def add_model_data_from_dict(section_dict, model_parameters):
    '''
    Die Sektionsweiße definierten Querschnittswerte werden so an das Model gegebn.
    NOTE/TODO: das ist hier alles ein bisschen kompliziert und eventuell verwirrend.
    Es gibt keine Intervale. Jeder x-Koordinate sind verschiedene QS werte zu geordnet da sich diese
    über die Höhe kontinuierlich verändern.
    - Knoten Anzahl wird an die Anzahl an Ebenen angepasst
    - Koordinaten definition wird hier an den Beam angepasst
    - Jedem Interval wird der Mittelwert der jeweiligen Querschnittswerte zugeordnet 
        (dieser kommt schon aus dem QS dictionary )
    '''
    model_parameters_updated = copy.deepcopy(model_parameters)
    
    model_parameters_updated['intervals'] = []

    model_parameters_updated['lz_total_beam'] = section_dict['section_absolute_heights'][-1]

    # TODO von section auf ebenen umstellen deutsch
    for section in range(section_dict['n_sections']):

        # 2D
        section_werte = {
            'bounds':[section_dict['section_absolute_heights'][section], section_dict['section_absolute_heights'][section+1]],
            'area': [section_dict['A'][section], section_dict['A'][section+1]],
            'Iy':[section_dict['Iy'][section], section_dict['Iy'][section+1]], 
            'D':[section_dict['d_achse'][section], section_dict['d_achse'][section+1]]
            }

        if model_parameters['dimension'] == '3D':
            # Aufgrund der Symmetrie eh das selbe
            section_werte['Iz'] = [section_dict['Iz'][section], section_dict['Iz'][section+1]]
                
        model_parameters_updated['intervals'].append(section_werte)

        if section == section_dict['n_ebenen']:
            model_parameters_updated['intervals']['bounds'][-1] = 'End'

    return model_parameters_updated

def initialize_empty_dict(dict_name, key_to_add):

    if key_to_add not in dict_name:
        dict_name[key_to_add] = {}
        return dict_name

    else:
        dict_name

def collect_max_ausn(results_df, max_ausnutzung:dict, lastfall):
    '''
    sucht in dem results_df in der dritten ebene (nach nabenhöhe und QS_label) nach variablen die die Ausnutzung ausgeben 
    diese updaten das max_ausnutzungsdict
    '''
    if not max_ausnutzung:
        lastfall0 = True
        for key in ['Wert', 'at_Höhe [m]', 'Ausn. max', 'Lastfall']:
            max_ausnutzung[key] = []
    else:
        lastfall0 = False

    df = results_df[lastfall]
    cols = df.columns
    for var in cols:
        if 'Ausn.' in var[-1]:
            col = df.loc[:,(var[0], var[1], var[2])].values
            h = df.loc[:,(var[0], var[1], 'Höhe [m]')].values
            
            if lastfall0:
                max_ausnutzung['Wert'].append(var[-1])
                max_ausnutzung['Ausn. max'].append(max(col))
                max_ausnutzung['at_Höhe [m]'].append(h[np.where(col == max(col))][0])
                max_ausnutzung['Lastfall'].append(lastfall)
            else:
                row = max_ausnutzung['Wert'].index(var[-1])
                if max_ausnutzung['Ausn. max'][row] < max(col):
                    max_ausnutzung['Ausn. max'][row] = max(col)
                    max_ausnutzung['at_Höhe [m]'][row] = h[np.where(col == max(col))][0]
                    max_ausnutzung['Lastfall'][row] = lastfall


    return max_ausnutzung

#_____________________OPENFAST / RFEM_______________________________________

def convert_coordinate_system_and_consider_einwirkungsdauer(loads:dict, n_nodes, gamma_g = 1.35, kopf_masse = 0, dimension:str = '2D'):
    '''
    Design Lasten aus IEA, OpenFAST in die Richtung vom Balken konvertieren
    Sortiert die Lasten nach einwirkungsdauer: Annahme -> Vertikallast = ständig die anderen kurz
    Dauer 'egal' gibt alle lasten zurück
    gamma_g: sicherheitsbeiwert für gewichtskraft wird aus den ständig wirkenden eigengewichten für die dauer "spannkraft" rausgerechnet 
            da Druckspannungen günstig beid er Berechnung der erforderlichen Sapnnkraft wirken
    kopf_masse: design gewichtskraft der Analgen masse wird für aufteilung in ständig und kurz rausgerechnet 
    TODO manueller Vorzeichen Wechsel des Moments ist in der Theorie falsch -> Achtung Lasten sind SGR also passt das alles?
    '''
    einwirkungsdauer = {'ständig':{'Fx':0, 'Fz':1, 'Fy':0, 'Mx':0, 'My':0, 'Mz':0}, 
                        'kurz':{'Fx':1, 'Fz':1, 'Fy':1, 'Mx':1, 'My':1, 'Mz':1}, 
                        'egal':{'Fx':1, 'Fz':1, 'Fy':1, 'Mx':1, 'My':1, 'Mz':1},
                        'spannkraft':{'Fx':1, 'Fy':1, 'Fz':1, 'Mx':1, 'My':1, 'Mz':1}}

    loads_beam = {'ständig':{}, 'kurz':{}, 'egal':{},'spannkraft':{}}

    converter_beam_fast = {'Fy':'Fx','Fx':'Fz', 'Fz':'Fy', 'Mx':'Mz', 'My':'Mx', 'Mz':'My'}

    # TODO rein theoretisch ist dieser Vorzeichen wechsel hier falsch -> Lasten sind SGR ?! 
    sign_beam_fast = {'Fx':1,'Fz':1, 'Fy':1, 'Mz':1 ,'Mx': 1, 'My':-1}
    for dauer in einwirkungsdauer:
        for direction in GD.FORCES[dimension]:
            loads_beam[dauer][direction] = np.zeros(n_nodes)
            factor = einwirkungsdauer[dauer][direction]
            kopflast_IEA = loads[direction] # Last IEA im beam cosyconverter_beam_fast[direction]

            if direction == 'Fz' and (dauer == 'ständig' or dauer == 'spannkraft'):
                # die ständige Last in Turm richtung ist nur die Kopfmasse
                kopflast_IEA = kopf_masse
            if direction == 'Fz' and dauer == 'kurz':
                # Die kurzzeitig wirkende Kraft in Turmrichtung ist die differenz der gemessenen zur kopfmasse
                kopflast_IEA -= kopf_masse

            loads_beam[dauer][direction][-1] += kopflast_IEA * factor * sign_beam_fast[direction] 

    # Für die Berechnung der Spannkraft -> Eigengewicht ohne sicherheitsbeiwert
    loads_beam['spannkraft']['Fz'] /= gamma_g #ohne sicherheitsbeiwert

    return loads_beam

def teilsicherheitsbeiwert_gravity_IEC(DLC, F_gravity, Fk):
    '''
    nach IEC 61400-1 Kapitel 7.6.2.2
    '''
    if DLC == 1.1:
        phi = 0.15
    else:
        phi = 0.25

    if abs(F_gravity) > abs(Fk):
        zeta = 0
    else:
        zeta = 1 - abs(F_gravity/Fk)

    gamma_f = 1.1 + phi*zeta**2
    return gamma_f


#____________________  ALLGEMEINES __________________________________________________

def convert_for_latex(string):
    l = list(string.split('_'))
    label = r'${}$'.format(l[0])
    for i in range(1,len(l)):
        label += r'$_{{{}}}$'.format(l[i])

    return label

def unit_conversion(unit_in:str, unit_out:str) -> float:
    '''
    gibt faktor zurück um einen Wert der Einheit unit_in in einen der Einheit unit_out umzurechnen
    '''
    converter = {'N/mm²':{'N/cm²':1E+02, 'N/m²':1E+06, 'kN/mm²':1E-03,'kN/m²':1E+03,'MPa':1},
                'N/m²':{'N/mm²':1E-06,'N/cm²':1E-04,'kN/mm²':1E-09,'kN/m²':1E-03,'Pa':1, 'MPa':1E-06},
                'N/cm²':{'N/mm²':1E-02},
                'kN/m²':{'N/mm²':1E-03,'N/cm²':1E-01,'kN/mm²':1E-06,'kN/m²':1,'MPa':1E-03},
                'N/m':{'kN/m':1E-03, 'MN/m':1E-06},
                'Nm/m':{'kNm/m':1E-03, 'MNm/m':1E-06},
                'N':{'kN':1E-03, 'MN':1E-06},
                'kN':{'N':1E+03, 'MN':1E-03},
                'MN':{'N':1E+06, 'kN':1E+03},
                'Nm':{'kNm':1E-03, 'MNm':1E-06},
                'kNm':{'Nm':1E+03, 'MNm':1E-03},
                'MNm':{'Nm':1E+06, 'kNm':1E+03},
                'm':{'cm':1E+02, 'mm':1E+03},
                'cm':{'m':1E-02, 'mm':1E+01},
                'mm':{'m':1E-02, 'cm':1E-01},
                'kWh':{'MWh':1/1000}}

    return converter[unit_in][unit_out]

def scale_dict_of_dicts(dd:dict, faktor):

    for k, subdict in dd.items():
        for k1 in subdict.keys():
            dd[k][k1] *= faktor
    
    return dd


#________________________ EXCEL STUFF ________________________

def get_excel_column_names():
    xl_cols = list(string.ascii_uppercase)
    ABC = list(string.ascii_uppercase)

    for char in list(string.ascii_uppercase):
        next_level = [char + abc for abc in ABC]
        xl_cols.extend(next_level)
    return xl_cols

EXCEL_COLUMNS = get_excel_column_names()

def excel_cell_name_to_row_col(cell_name):

    row = int(cell_name[1:]) - 1 
    col = string.ascii_uppercase.index(cell_name[0])
    return row, col

def read_xl_column(xl_worksheet, start_cell:str = None, start_row:int = None, start_col:int = None, end_row:int=None, end_cell:str = None):
    '''
    reading excel colum startin from start_row, start_col ODER start_cell als Excel Zellen name
    start_row = Excel Zeilen Zahl (wird intern dann angepasst auf das 0 zählsystem von xlwings)
    start_col = A entspricht 0 
    end_row_optional: wenn nicht gegebn wird spalte bis zur nächsten leeren Zeile gelesen NOTE klappt nicht immer 
    '''
    if start_cell:
        if start_col or start_row:
            raise Exception('entweder Start Zelle als String oder Zeile und Spalte als nummern geben')

        start_row, start_col = excel_cell_name_to_row_col(start_cell)

    else:
        start_row -=1
         
    if not end_row and not end_cell:
        end_row = xl_worksheet.range(start_row, start_col).end('down').row

    if end_cell:
        if end_row:
            raise Exception('End Zelle als String oder Zeilen nummer angeben')

        end_row, end_col = excel_cell_name_to_row_col(end_cell)
    
    column_values = xl_worksheet[start_row:end_row, start_col].value
    return column_values

def read_xl_row(xl_worksheet, start_row, start_col):
    '''
    reading excel row startin from start_row, start_col
    '''
    #sheet["A1"].expand("right").last_cell.column
    end_col = xl_worksheet[start_row, start_col].expand('right').last_cell.column
    row_values = xl_worksheet[start_row, start_col:end_col].value
    return row_values

def write_to_excel(worksheet, start_cell, values, header = ['Hallo'], is_row = False, is_column = True):
    '''
    Werte in excel blatt schreibenbeginnend ab start cell
    worksheet: ein mit xlwings geöffnetes Excel Arbeitsblatt
    start_cell kann als Excel notation ('A1') oder rwo und column liste kommen 
    '''
    ws = worksheet

    for i in values:
        if isinstance(i, list):
            i = np.asarray(list)

    if isinstance(values, list):
        values = np.asarray(values)

    # convert values to a pandas dataframe -> das kann sehr gut in Excel verwendet werden
    if not isinstance(values, pd.DataFrame):
        values = values.transpose()
        df = pd.DataFrame(values, columns=header)

    if isinstance(start_cell, str):
        ws.range(start_cell).value = df
    else:
        ws.range(start_cell[0],start_cell[1]).value = df

    print ('saved', header, 'in', ws.name)

def get_simulation_history_id(case_name):

    wb = xl.Book(os_join(*['..','1_Simulations','simulation_history.xlsx']))
    ws = wb.sheets['parameter']
    header = read_xl_row(ws, 1, 2)
    col = header.index('outfile name')+2
    all_cases = ws[0:300, col].value
    row = all_cases.index(case_name)
    case_id = int(ws[row,1].value)

    return case_id

def write_stats_to_excel(excel_file, worksheet, start_cell, statistics_to_write, stats_results, case_name):
    '''
    TODO variablen die noch nicht drinnen sind im Excel händsich hinzufügen oder hier automatisch
    '''
    wb = xl.Book(excel_file)
    ws = wb.sheets[worksheet]
    header = read_xl_row(ws, 4, 2)

    if ws[start_cell[0], start_cell[1]].value != None:
        weiter = input('start_cell ist schon voll ' + worksheet + ' ' +str(start_cell[0]) + ',' +str(start_cell[1]) + ' -> wenn y wirds überschrieben? [y]/[n] ')
        if weiter == 'y':
            pass
        else:
            raise Exception('start_cell neu wählen -> row = letzte beschriebene Reihe + 1')

    ws[start_cell[0], 0].value = case_name 
    
    # id der simulation history hinzufügen
    case_id = get_simulation_history_id(case_name)
    ws[start_cell[0], 1].value = case_id

    for var in header:
        col = start_cell[1] + header.index(var)
        value = stats_results[var][statistics_to_write]
        ws[start_cell[0], col].value = value 
    wb.save()

def zellen_groeße_formatieren(excel_file, worksheet:str, df:pd.DataFrame=None, cell_width:int=15, n_cols:int=1,  start_col:str = 'A', cell_height = None):
    ''' 
    excel_file: datei pfad
    Worksheet: sheet name
    df: der dataframe der geschrieben wird -> cell_width und n_cols wird angepasst wenn vorhanden NOTE funktioniert nicht für Multiindex df
    cell_width: breite der zell
    n_cols: anzahl der spalten die formatiert werden sollen 
    start_col: erste Spalte ab der das formatieren los gehen soll (als Excel Buschtabe oder nummer)
    '''
    from openpyxl import load_workbook

    wb = load_workbook(excel_file)
    ws = wb[worksheet]

    if isinstance(start_col, str):
        start_col = EXCEL_COLUMNS.index(start_col)

    if isinstance(df, pd.DataFrame):
        c =  df.columns
        b = [len(x) for x in df.columns]
        cell_width = max([len(x) for x in df.columns])
        n_cols = len(df.columns) + 1

    if start_col == n_cols:
        ws.column_dimensions[start_col].width = cell_width
    else:
        for i in range(start_col, n_cols):
            excel_col = EXCEL_COLUMNS[i]
            ws.column_dimensions[excel_col].width = cell_width
            
    wb.save(excel_file)
    wb.close()

def add_databar_color(excel_file, worksheet, columns):
    '''
    columns liste aus columns: ['B1:B12', ...]
    https://prosperocoder.com/posts/science-with-python/openpyxl-part-16-styling-with-databar/ 
    '''

    from openpyxl import load_workbook
    from openpyxl.formatting.rule import DataBarRule
    #from openpyxl.styles import colors

    wb = load_workbook(excel_file)
    ws = wb[worksheet]

    ROT = '00FF0000'
    rule = DataBarRule(start_type='num', 
                   start_value=0, 
                   end_type='num', 
                   end_value=1, 
                   color=ROT) # Rot:'00FF0000'

    for col in columns:
        ws.conditional_formatting.add(col, rule)
    
    wb.save(excel_file)
    wb.close()

def add_color_rule(excel_file, worksheet, columns, regel = 'lessThan', farbe = 'rot', value = 0):
    '''    
    '''
    from openpyxl import load_workbook, styles
    from openpyxl.formatting.rule import CellIsRule

    wb = load_workbook(excel_file)
    ws = wb[worksheet]

    farben = {'rot':'ffc7ce'}#'ffc7ce''00FF0000'

    color_fill = styles.PatternFill(start_color=farben[farbe], end_color=farben[farbe], fill_type='solid')

    rule = CellIsRule(operator=regel, formula=[str(value)], fill=color_fill) # Rot:'00FF0000'

    for col in columns:
        ws.conditional_formatting.add(col, rule)
    
    wb.save(excel_file)
    wb.close()


def zelle_beschriften(excel_file, worksheet, cell, wert, merge_from_to = None, zelle_färben = False):
    '''
    merge_from_to: 'A2:D2' Zellen verbinden
    '''
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    wb = load_workbook(excel_file)
    ws = wb[worksheet]

    if not isinstance(cell, list):
        cell = list(cell)

    for cell_i in cell:
        ws[cell_i].value = wert
        if zelle_färben:
            ws[cell_i].fill = PatternFill(fill_type='solid', fgColor='FFD3D3D3')

    if merge_from_to:
        ws.merge_cells(merge_from_to)
    wb.save(excel_file)
    wb.close()

def spalte_beschriften(excel_file, worksheet, start_cell, wert:list, merge_from_to = None, zelle_färben = False):
    '''
    merge_from_to: 'A2:D2' Zellen verbinden
    start_cell in Excel format: A1
    '''
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    wb = load_workbook(excel_file)
    ws = wb[worksheet]

    for i in wert:
        i = i.translate(None, '-+=')
        ws[start_cell].value = i
        if zelle_färben:
            ws[start_cell].fill = PatternFill(fill_type='solid', fgColor='FFD3D3D3')
        
        start_cell = list(start_cell)[0] + str(int(list(start_cell)[1])+1)

    if merge_from_to:
        ws.merge_cells(merge_from_to)
    wb.save(excel_file)
    wb.close()

def get_spalten_ausnutzung(df, df_header, start_row):

    n_rows = df.shape[0]
    end_row = start_row + n_rows + len(df_header) +1
    cols, nth_col = [], []
    # n-te Spalte je Querschnitt 
    for i, val in enumerate(df_header[-1]):
        if 'Ausnutzung' in val or 'Ausn.' in val:
            nth_col.append(i+1) # index vom df ist erst spalte
    
    # alle Spalten die gesamt
    for j in nth_col:
        for i in range(df.columns.levshape[0]*df.columns.levshape[1]):
            cols.append(j + i*df.columns.levshape[2])

    xl_col = [EXCEL_COLUMNS[i] for i in cols]

    cols_rows = [i + str(start_row) + ':' + i+ str(end_row) for i in xl_col]

    return cols_rows

def get_spalten_nach_name(df, df_header, start_row, name):

    #start_row = len (df_header)
    n_rows = df.shape[0]
    start_row += len(df_header) + 1
    end_row = start_row + n_rows  
    cols, nth_col = [], []
    # n-te Spalte je Querschnitt 
    for i, val in enumerate(df_header[-1]):
        if name in val:
            nth_col.append(i+1) # index vom df ist erst spalte
    
    # alle Spalten die gesamt
    for j in nth_col:
        depth = df.columns.levshape[0]*df.columns.levshape[1]*df.columns.levshape[2]
        for i in range(depth):
            cols.append(j + i*df.columns.levshape[3])

    xl_col = [EXCEL_COLUMNS[i] for i in cols]

    cols_rows = [i + str(start_row) + ':' + i+ str(end_row) for i in xl_col]

    return cols_rows


def save_dataframes_to_pkl(data:dict, knoten_typ:str, destination:list=['output','Berechnungs_Ergebnisse_dataframes']):
    '''
    keys der data sollten die NAmen der zu speichernden Datei sein
    knoten_typ: zeigt an an welchen der Knoten die Ergebnisse ausgegeben sind
    '''
    from datetime import datetime
    now = datetime.now()
    dt_string = now.strftime("%Y%m%d")[2:]

    for name, df in data.items():
        destination.append(dt_string + '_' + name + '_' + knoten_typ + '.pkl')
        dest_file = os_join(*destination)
        df.to_pickle(dest_file)
        del destination[-1]

    

#______________________ BERECHNUNGEN __________________________

def twr_poly_eval(x, coeffs):
    ''' beginnt erst am ^2 geht bis ^6 und c0 = 0'''
    power = 2
    y = np.zeros(x.shape)
    for c_i, c in enumerate(coeffs):
        y += c * x ** (power+c_i)

    return y

def collect_mode_shape_params(pyfast_file):

    mode_shapes_poly_coeffs = {'fore-aft':[[],[]],'side-side':[[],[]]}
    for label in pyfast_file.labels:
        if 'TwFAM' in label:
            if '1Sh' in label:
                mode_shapes_poly_coeffs['fore-aft'][0].append(pyfast_file[label])
            else:
                mode_shapes_poly_coeffs['fore-aft'][1].append(pyfast_file[label])
        if 'TwSSM' in label:
            if '1Sh' in label:
                mode_shapes_poly_coeffs['side-side'][0].append(pyfast_file[label])
            else:
                mode_shapes_poly_coeffs['side-side'][1].append(pyfast_file[label])
    return mode_shapes_poly_coeffs

def compute_sectional_mean(parameter):
        
    if not isinstance(parameter, list):
        parameter = list(parameter)
    sectional_sum = [j+i for i, j in zip(parameter[:-1], parameter[1:])]
    return np.array(sectional_sum)/2

def get_d_achse_lagen(d_achse, lagenaufbau):
    '''
    d_achse: Achs-Durchmesser des gesamtquerschnitts
    lagenaufbau: dict wie lagenaufbau
    fügt 'di' der jeweiligen Lage hinzu
    '''
    mitte = int(len(lagenaufbau)/2) 

    for i, lage in enumerate(lagenaufbau):
        if i == mitte:
            lage['di'] = d_achse
            continue
        # TODO für Mittellage sollte auch 11 rauskommen --> checke das alles
        if i <= mitte:
            ganze_lagen = sum([l['ti'] for l in lagenaufbau][i+1:mitte])
            lage_selber = lage['ti']/2 
            mittellage =lagenaufbau[mitte]['ti']/2
            di = d_achse - ganze_lagen - lage_selber - mittellage
        else:
            di = d_achse + sum([l['ti'] for l in lagenaufbau][mitte+1:i]) + lage['ti']/2 + lagenaufbau[mitte]['ti']/2
        lage['di'] = di

# # DATEN MANGAMENT

def save_specific_columns_from_fast_output(cols_to_keep:list, fast_output:pd.DataFrame, fname:str):
    '''
    cols_to_keep: liste an output variablen in der FAST out datei die gespeichert werden soll
    fast_output: mit FAST_inputoutput Klasse gelesens und in dataframe konvertiertes out file von FAST
    fname: name/pfad der datei die als pkl gespeichert weren soll
    '''
    cols_to_keep.append('Time_[s]')
    fast_output.loc[:, cols_to_keep].to_pickle(fname)

    print ('\nsaved', cols_to_keep, 'in', fname)
    



