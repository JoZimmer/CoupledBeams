# from math import log10
# from source.utilities import utilities as utils
# from inputs import holz
# import matplotlib.pyplot as plt
import numpy as np
# import inputs.DIN_Windlasten as windDIN
import pandas as pd
from os.path import join as os_join

from openpyxl import formatting, styles, Workbook

wb = Workbook()
ws = wb.active

red_color = 'ffc7ce'
red_color_font = '9c0103'

red_font = styles.Font(size=14, bold=True, color=red_color_font)
red_fill = styles.PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')

for row in range(1,10):            
    ws.cell(row=row, column=1, value=row-5)
    ws.cell(row=row, column=2, value=row-5)

rule = formatting.rule.CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)

#ws.conditional_formatting.add('A1:A10', formatting.rule.CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font))
ws.conditional_formatting.add('B1:B10', rule)
wb.save("test.xlsx")
    
